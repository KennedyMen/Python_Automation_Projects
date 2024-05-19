from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook(
    '/Users/033103kennedymensah/Automation/Excel_Setup/pivot_table.xlsx')
sheet = wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column)
print(max_column)
print(min_row)
print(max_row)
# --------------------------------------------------------------------------

# Stanard Pivot Table for any setup
BarChart = BarChart()

Data = Reference(sheet, min_row=min_row, max_row=max_row,
                 min_col=min_column+1, max_col=max_column)
catagories = Reference(sheet, min_row=min_row+1,
                       max_row=max_row, min_col=min_column, max_col=min_column)


BarChart.add_data(Data, titles_from_data=True)
BarChart.set_categories(catagories)
sheet.add_chart(BarChart, "B12")
BarChart.title = "Sales by Product Line"
BarChart.style = 5
wb.save('barchart.xlsx')
