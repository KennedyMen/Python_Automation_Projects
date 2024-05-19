from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook(
    '/Users/033103kennedymensah/Automation/Excel_Setup/barchart.xlsx')
sheet = wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

sheet['B8'] = '=SUM(B6:B7)'
sheet['B8'].style = 'Currency'

wb.save('report.xlsx')
